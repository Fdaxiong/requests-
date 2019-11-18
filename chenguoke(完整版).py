import requests
from lxml import etree
from pprint import pprint
import json
from openpyxl import Workbook

url = "https://www.guokr.com/ask/hottest/"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36"}

response = requests.get(url, headers=headers)
response = response.content.decode()
html = etree.HTML(response)
answer_list = []
ul_list = html.xpath("//ul[@class='ask-list-cp']")
data_dict = {}
data_dict["title"] = []
data_dict["href"] = []
while True:
    answer = {}
    response = requests.get(url, headers=headers)
    response = response.content.decode()
    html = etree.HTML(response)
    ul_list = html.xpath("//ul[@class='ask-list-cp']/li")
    for li in ul_list:
        text = (li.xpath(".//h2/a/text()"))[0] if len(li.xpath(".//h2/a/text()")) >= 1 else None
        # print("text",text)
        data_dict["title"].append(text)
        href = (li.xpath(".//h2/a/@href"))[0] if len(li.xpath(".//h2/a/@href")) >= 1 else None
        # print("href", href)
        data_dict["href"].append(href)
    print(data_dict)
    # print(answer_list)
    next_next = html.xpath("//ul[@class='gpages']/li/a[text()='下一页']/@href")
    str = ""
    url = "https://www.guokr.com" + str.join(next_next)
    # print(url)
    if url == "https://www.guokr.com":
        break

# 爬取得结果给data_data
# data_data = answer
# pprint(type(data_data))
# with open("guoke.txt","w",encoding="utf-8") as f:
#     f.write(json.dumps(data_data,ensure_ascii=False,indent=2))

# data_data_b = {'href': ['https://www.guokr.com/question/654665/',
#                       'https://www.guokr.com/question/666807/',
#                       'https://www.guokr.com/question/670597/',
#                       'https://www.guokr.com/question/669759/',
#                       'https://www.guokr.com/question/678361/',
#                       'https://www.guokr.com/question/102172/',
#                       'https://www.guokr.com/question/653285/',
#                       'https://www.guokr.com/question/375294/',
#                       'https://www.guokr.com/question/666053/',
#                       'https://www.guokr.com/question/670934/',
#                       'https://www.guokr.com/question/648157/',
#                       'https://www.guokr.com/question/484734/',
#                       'https://www.guokr.com/question/339467/',
#                       'https://www.guokr.com/question/313334/',
#                       'https://www.guokr.com/question/229234/',
#                       'https://www.guokr.com/question/528807/'],
#              'text': ['在中国，如果人人有枪，是否会多一些人与人之间的相互敬畏与尊重？',
#                       '为何好多人都说中药不科学？不治病？它是怎么个工作原理？',
#                       '在原料中的物种本来就没有转基因品种的情况下标注非转基因是否涉及不当竞争？',
#                       '为什么人类混血的好看，其他动物都是纯血的好看',
#                       '什么是“清真食品”？',
#                       '果壳网的运营资金从何而来呢？',
#                       '妈妈没收了鼠标后该怎么玩电脑？',
#                       '为什么我向果壳问答提问从来没有得到回答呢？希望这次能有答案啊啊啊啊！',
#                       '为什么小狗的叫声是“汪汪汪”，而不是“喵喵喵”或者其他什么声音？是什么原因造成的？',
#                       '《救命饮食》发行很久了，里面有许多科学实验证明人类吃肉会致病，为什么学校的营养书里还是建议人吃肉?是因为无知?还是因为不想麻烦?为什么对大众的身体健康这么不负责任?',
#                       '用洗衣粉洗头，会水解~~~？',
#                       '如果一斤糖和一斤盐放在一个碗里，怎么把它们完全分离出来？',
#                       '化学药品的味道都是先驱们去尝的么？',
#                       '猪一天需要睡多少小时？',
#                       '根据西瓜的形状、颜色和纹路来挑西瓜有科学依据吗？',
#                       '如何评价“如何评价XXXX？”这样的问题？']}


try:
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 100
    ws["A1"] = "问题"
    ws["B1"] = "链接"
    for i in data_dict["title"]:
        print(type(i))
        ws.append([i.encode("utf-8"), data_dict["href"][data_dict["title"].index(i)].encode("utf-8")])
    wb.save("果壳问答.xlsx")
except PermissionError:
    print("你没有关闭Excel,又去执行读写操作,关闭Excel重试")
print("***************")
pprint(data_dict)