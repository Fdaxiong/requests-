from lxml import etree
from pprint import pprint
import requests
import collections
from openpyxl import Workbook



class GuoKe(object):
    # 初始化(前期准备)
    def __init__(self):
        self.url = "https://www.guokr.com/ask/highlight/"
        self.headers = {
            "Host": "www.guokr.com",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36"
        }
        self.prefix = "https://www.guokr.com"
        self.data_list = []
        self.title_length = 0
        self.link_length = 0
        self.comment_length = 0

    def run(self):
        i = 0
        while i<1:
            # 实现总体流程控制
            response = self.get_response(self.url)
            # 提取数据
            self.parse_content(response)
            # 下一页的内容
            next_page = self.get_next_url(response)
            if not next_page:
                print(self.data_list)
                break
            else:
                self.url = next_page
                i += 1
        self.save_data()

    def get_response(self,url):
        resp = requests.get(url,headers=self.headers)
        return resp.content.decode()

    def parse_content(self, response):
        html = etree.HTML(response)
        li_list = html.xpath('//ul[@class="ask-list-cp"]/li')
        for li in li_list:
            item = collections.OrderedDict()
            item["title"] = li.xpath('./div[@class="ask-list-detials"]/h2/a/text()')[0] if len(li.xpath('./div[@class="ask-list-detials"]/h2/a/text()'))>0 else None
            item["link_url"] = li.xpath('./div[@class="ask-list-detials"]/h2/a/@href')[0] if len(li.xpath('./div[@class="ask-list-detials"]/h2/a/@href'))>0 else None
            response = self.get_response(item["link_url"])
            comments = self.get_detail_comment(response)
            item["comments"] = comments
            self.data_list.append(item)

    def get_detail_comment(self, response):
        html = etree.HTML(response)
        comments = html.xpath('//div[@class="answer-txt answerTxt gbbcode-content"]//p/text()')
        return comments

    def get_next_url(self,res):
        html = etree.HTML(res)
        nex_page = html.xpath('//a[text()="下一页"]/@href')[0] if len(html.xpath('//a[text()="下一页"]/@href'))>0 else None
        next_url = self.prefix + nex_page
        return next_url

    def update_length(self,data,ws):
        title_length = len(data["title"].encode("utf-8"))
        link_length = len(data["link_url"].encode("utf-8"))
        comment_length = len(data["comments"][0].encode("utf-8"))
        if title_length > self.title_length:
            self.title_length = title_length
        if link_length > self.link_length:
            self.link_length = link_length
        if comment_length > self.comment_length:
            self.comment_length = comment_length
        ws.column_dimensions['A'].width = self.title_length
        ws.column_dimensions['B'].width = self.link_length
        ws.column_dimensions['C'].width = self.comment_length

    def save_data(self):
        wb = Workbook()
        # 得到一个sheet 表
        ws = wb.active  # 获取第一个sheet
        ws['A1'] = "提问内容"
        ws['B1'] = "提问链接"
        ws['C1'] = "最热评论"
        for data in self.data_list:
            self.update_length(data, ws)
            ws.append((data["title"].encode("utf-8"),data["link_url"].encode("utf-8"),data["comments"][0].encode("utf-8")))
        wb.save("guoke.xlsx")

if __name__ == '__main__':
    guoke = GuoKe()
    guoke.run()